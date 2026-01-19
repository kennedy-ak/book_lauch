from django.contrib import admin
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Visitor


@admin.register(Visitor)
class VisitorAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone_number', 'location', 'created_at']
    list_filter = ['created_at', 'location']
    search_fields = ['name', 'email', 'phone_number', 'location']
    readonly_fields = ['created_at']
    ordering = ['-created_at']
    
    actions = ['export_as_csv', 'export_as_excel']
    
    def export_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="visitors.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Phone Number', 'Email', 'Location', 'Created At'])
        
        for visitor in queryset:
            writer.writerow([
                visitor.name,
                visitor.phone_number,
                visitor.email,
                visitor.location,
                visitor.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    export_as_csv.short_description = "Export selected to CSV"
    
    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="visitors.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Visitors'
        
        # Write headers
        headers = ['Name', 'Phone Number', 'Email', 'Location', 'Created At']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, visitor in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=visitor.name)
            ws.cell(row=row_num, column=2, value=visitor.phone_number)
            ws.cell(row=row_num, column=3, value=visitor.email)
            ws.cell(row=row_num, column=4, value=visitor.location)
            ws.cell(row=row_num, column=5, value=visitor.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
        return response
    
    export_as_excel.short_description = "Export selected to Excel"
