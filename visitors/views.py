from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from .models import Visitor
from .forms import VisitorForm
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils import timezone


def admin_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is None:
                messages.error(request, 'Invalid username or password.')
            else:
                login(request, user)
                if user.is_staff:
                    return redirect('dashboard')
                else:
                    messages.error(request, 'You do not have staff access.')
                    logout(request)
                    return redirect('register_visitor')
    else:
        form = AuthenticationForm()
    
    return render(request, 'visitors/admin_login.html', {'form': form})


def admin_logout(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('register_visitor')


def register_visitor(request):
    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thank you for registering! Your information has been saved.')
            return redirect('thank_you')
    else:
        form = VisitorForm()
    
    return render(request, 'visitors/register.html', {'form': form})


def thank_you(request):
    return render(request, 'visitors/thank_you.html')


def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="visitors.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Phone Number', 'Email', 'Location', 'Created At'])
    
    visitors = Visitor.objects.all()
    for visitor in visitors:
        writer.writerow([
            visitor.name,
            visitor.phone_number,
            visitor.email,
            visitor.location,
            visitor.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="visitors.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Visitors'
    
    # Write headers
    headers = ['Name', 'Phone Number', 'Email', 'Location', 'Created At']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    visitors = Visitor.objects.all()
    for row_num, visitor in enumerate(visitors, 2):
        ws.cell(row=row_num, column=1, value=visitor.name)
        ws.cell(row=row_num, column=2, value=visitor.phone_number)
        ws.cell(row=row_num, column=3, value=visitor.email)
        ws.cell(row=row_num, column=4, value=visitor.location)
        ws.cell(row=row_num, column=5, value=visitor.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(response)
    return response


@staff_member_required
def dashboard(request):
    visitors = Visitor.objects.all()
    total_visitors = visitors.count()
    
    context = {
        'visitors': visitors,
        'total_visitors': total_visitors,
    }
    return render(request, 'visitors/dashboard.html', context)


@staff_member_required
def dashboard_export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="visitors.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Phone Number', 'Email', 'Location', 'Created At'])
    
    visitors = Visitor.objects.all()
    for visitor in visitors:
        writer.writerow([
            visitor.name,
            visitor.phone_number,
            visitor.email,
            visitor.location,
            visitor.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


@staff_member_required
def dashboard_export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="visitors.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Visitors'
    
    # Write headers
    headers = ['Name', 'Phone Number', 'Email', 'Location', 'Created At']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    visitors = Visitor.objects.all()
    for row_num, visitor in enumerate(visitors, 2):
        ws.cell(row=row_num, column=1, value=visitor.name)
        ws.cell(row=row_num, column=2, value=visitor.phone_number)
        ws.cell(row=row_num, column=3, value=visitor.email)
        ws.cell(row=row_num, column=4, value=visitor.location)
        ws.cell(row=row_num, column=5, value=visitor.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(response)
    return response
